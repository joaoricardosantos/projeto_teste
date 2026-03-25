import threading
import uuid
import os
import tempfile
from datetime import datetime, timedelta
from typing import List, Optional

import requests
from django.http import HttpResponse, FileResponse
from ninja import Router, Schema
from ninja.errors import HttpError
from pydantic import UUID4, EmailStr

from core.auth import JWTAuth
from core.models import User
from core.superlogica import gerar_relatorio_inadimplentes, buscar_unidades

admin_router = Router(auth=JWTAuth())

def _is_admin(user):
    return user.is_staff or user.is_superuser

def _require_admin(user):
    if not _is_admin(user):
        raise HttpError(403, "Admin_privileges_required")

def _require_approved(user):
    if not user.is_approved:
        raise HttpError(403, "Account_not_approved")

# ── Job store em memória ──────────────────────────────────────────────────────
_JOBS: dict = {}
_JOBS_LOCK = threading.Lock()

# ── Cache do dashboard ────────────────────────────────────────────────────────
_DASHBOARD_CACHE: dict = {}
_CACHE_LOCK = threading.Lock()
_CACHE_TTL_MINUTES = 30

def _proximas_8h():
    agora = datetime.now()
    proximas = agora.replace(hour=8, minute=0, second=0, microsecond=0)
    if agora >= proximas:
        proximas += timedelta(days=1)
    return proximas


def _run_job(job_id: str, id_condominio, data_posicao, data_inicio=None, ordenar_desc=False):
    with _JOBS_LOCK:
        _JOBS[job_id]["status"] = "running"
    try:
        try:
            content, filename = gerar_relatorio_inadimplentes(
                id_condominio=id_condominio,
                data_posicao=data_posicao,
                data_inicio=data_inicio,
                ordenar_desc=ordenar_desc,
            )
        except TypeError:
            content, filename = gerar_relatorio_inadimplentes(
                id_condominio=id_condominio,
                data_posicao=data_posicao,
            )
        if not content:
            with _JOBS_LOCK:
                _JOBS[job_id]["status"] = "empty"
            return

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(content)
        tmp.close()

        with _JOBS_LOCK:
            _JOBS[job_id]["status"]   = "done"
            _JOBS[job_id]["file"]     = tmp.name
            _JOBS[job_id]["filename"] = filename

    except Exception as e:
        with _JOBS_LOCK:
            _JOBS[job_id]["status"] = "error"
            _JOBS[job_id]["error"]  = str(e)


# ── Schemas ───────────────────────────────────────────────────────────────────
class UserApprovalIn(Schema):
    user_id: UUID4
    is_approved: bool

class UserCreateIn(Schema):
    name: str
    email: EmailStr
    password: str

class UserOut(Schema):
    id: UUID4
    name: str
    email: str
    is_approved: bool
    is_active: bool
    is_staff: bool
    is_superuser: bool
    is_juridico: bool

class AdminRoleIn(Schema):
    user_id: UUID4
    make_admin: bool

class UserDeleteIn(Schema):
    user_id: UUID4

class UserJuridicoIn(Schema):
    user_id: UUID4
    is_juridico: bool


# ── Endpoints de usuários ─────────────────────────────────────────────────────
@admin_router.get("/users", response=List[UserOut])
def list_users(request):
    _require_admin(request.auth)
    return User.objects.all().order_by("-created_at")


@admin_router.post("/approve-user", response={200: dict})
def approve_user(request, payload: UserApprovalIn):
    _require_admin(request.auth)
    try:
        user = User.objects.get(id=payload.user_id)
        user.is_approved = payload.is_approved
        user.save()
        return 200, {"message": "User_approval_status_updated"}
    except User.DoesNotExist:
        raise HttpError(404, "User_not_found")


@admin_router.post("/create-user", response={201: dict})
def create_user(request, payload: UserCreateIn):
    _require_admin(request.auth)
    if User.objects.filter(email=payload.email).exists():
        raise HttpError(400, "Email_already_registered")
    User.objects.create_user(
        email=payload.email,
        password=payload.password,
        name=payload.name,
    )
    return 201, {"message": "User_created_successfully_pending_approval"}


@admin_router.post("/set-admin", response={200: dict})
def set_admin_role(request, payload: AdminRoleIn):
    _require_admin(request.auth)
    try:
        user = User.objects.get(id=payload.user_id)
    except User.DoesNotExist:
        raise HttpError(404, "User_not_found")
    user.is_staff = payload.make_admin
    user.is_superuser = payload.make_admin
    user.save()
    return 200, {"message": "User_admin_role_updated"}


@admin_router.post("/set-juridico", response={200: dict})
def set_juridico_role(request, payload: UserJuridicoIn):
    _require_admin(request.auth)
    try:
        user = User.objects.get(id=payload.user_id)
    except User.DoesNotExist:
        raise HttpError(404, "User_not_found")
    user.is_juridico = payload.is_juridico
    user.save()
    return 200, {"message": "User_juridico_role_updated"}


@admin_router.delete("/delete-user", response={200: dict})
def delete_user(request, payload: UserDeleteIn):
    _require_admin(request.auth)
    if str(request.auth.id) == str(payload.user_id):
        raise HttpError(400, "Cannot_delete_own_account")
    try:
        user = User.objects.get(id=payload.user_id)
        user.delete()
        return 200, {"message": "User_deleted_successfully"}
    except User.DoesNotExist:
        raise HttpError(404, "User_not_found")


# ── Endpoints de relatório (background job) ───────────────────────────────────
@admin_router.post("/export-defaulters/start", response={202: dict})
def start_export(
    request,
    id_condominio: Optional[str] = None,
    data_posicao: Optional[str] = None,
    ultimos_5_anos: bool = False,
    ordenar_desc: bool = False,
):
    """
    Inicia a geração do relatório em background.
    id_condominio pode ser um único ID ou vários separados por vírgula.
    ultimos_5_anos=true filtra vencimentos dos últimos 5 anos.
    ordenar_desc=true ordena por Total decrescente.
    """
    _require_approved(request.auth)

    data_inicio = None
    if ultimos_5_anos:
        data_inicio = (datetime.now() - timedelta(days=5 * 365)).strftime("%d/%m/%Y")

    ids = None
    if id_condominio:
        partes = [p.strip() for p in id_condominio.split(",") if p.strip()]
        if len(partes) == 1:
            ids = int(partes[0])
        else:
            ids = [int(p) for p in partes]

    job_id = str(uuid.uuid4())
    with _JOBS_LOCK:
        _JOBS[job_id] = {"status": "pending", "file": None, "filename": None, "error": None}

    t = threading.Thread(
        target=_run_job,
        args=(job_id, ids, data_posicao, data_inicio, ordenar_desc),
        daemon=True,
    )
    t.start()

    return 202, {"job_id": job_id}


@admin_router.get("/export-defaulters/status/{job_id}", response={200: dict})
def job_status(request, job_id: str):
    _require_approved(request.auth)
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        raise HttpError(404, "Job_not_found")
    return 200, {
        "status":   job["status"],
        "filename": job.get("filename"),
        "error":    job.get("error"),
    }


@admin_router.get("/export-defaulters/download/{job_id}")
def download_export(request, job_id: str):
    _require_approved(request.auth)
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        raise HttpError(404, "Job_not_found")
    if job["status"] != "done":
        raise HttpError(400, "Job_not_ready")
    filepath = job["file"]
    filename = job["filename"]
    if not filepath or not os.path.exists(filepath):
        raise HttpError(404, "File_not_found")
    with open(filepath, "rb") as f:
        content = f.read()
    try:
        os.unlink(filepath)
    except Exception:
        pass
    with _JOBS_LOCK:
        _JOBS.pop(job_id, None)
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── Endpoint legado (compatibilidade) ─────────────────────────────────────────
@admin_router.get("/export-defaulters")
def export_defaulters(
    request,
    id_condominio: Optional[int] = None,
    data_posicao: Optional[str] = None,
):
    _require_approved(request.auth)
    if not id_condominio:
        raise HttpError(400, "Use_async_endpoint_for_all_condominios")
    try:
        content, filename = gerar_relatorio_inadimplentes(
            id_condominio=id_condominio,
            data_posicao=data_posicao,
        )
    except requests.RequestException:
        raise HttpError(502, "External_service_unavailable")
    if not content:
        raise HttpError(204, "No_defaulters_found")
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── Limpar cache do dashboard ─────────────────────────────────────────────────
@admin_router.post("/dashboard/clear-cache", response={200: dict})
def clear_dashboard_cache(request):
    _require_approved(request.auth)
    with _CACHE_LOCK:
        _DASHBOARD_CACHE.clear()
    return 200, {"message": "Cache limpo com sucesso"}


# ── Endpoints de PDF (background job) ────────────────────────────────────────
@admin_router.post("/export-pdf/start", response={202: dict})
def start_export_pdf(
    request,
    id_condominio: Optional[str] = None,
    data_posicao: Optional[str] = None,
    ultimos_5_anos: bool = False,
    ordenar_desc: bool = False,
):
    _require_approved(request.auth)

    from core.superlogica import gerar_pdf_inadimplentes

    data_inicio = None
    if ultimos_5_anos:
        data_inicio = (datetime.now() - timedelta(days=5 * 365)).strftime("%d/%m/%Y")

    ids = None
    if id_condominio:
        partes = [p.strip() for p in id_condominio.split(",") if p.strip()]
        if len(partes) == 1:
            ids = int(partes[0])
        else:
            ids = [int(p) for p in partes]

    job_id = str(uuid.uuid4())
    with _JOBS_LOCK:
        _JOBS[job_id] = {"status": "pending", "file": None, "filename": None, "error": None}

    def _run():
        with _JOBS_LOCK:
            _JOBS[job_id]["status"] = "running"
        try:
            try:
                content, filename = gerar_pdf_inadimplentes(
                    id_condominio=ids,
                    data_posicao=data_posicao,
                    data_inicio=data_inicio,
                    ordenar_desc=ordenar_desc,
                )
            except TypeError:
                content, filename = gerar_pdf_inadimplentes(
                    id_condominio=ids,
                    data_posicao=data_posicao,
                )
            if not content:
                with _JOBS_LOCK:
                    _JOBS[job_id]["status"] = "empty"
                return
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            tmp.write(content)
            tmp.close()
            with _JOBS_LOCK:
                _JOBS[job_id]["status"]   = "done"
                _JOBS[job_id]["file"]     = tmp.name
                _JOBS[job_id]["filename"] = filename
        except Exception as e:
            with _JOBS_LOCK:
                _JOBS[job_id]["status"] = "error"
                _JOBS[job_id]["error"]  = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return 202, {"job_id": job_id}


@admin_router.get("/export-pdf/status/{job_id}", response={200: dict})
def pdf_job_status(request, job_id: str):
    _require_approved(request.auth)
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        raise HttpError(404, "Job_not_found")
    return 200, {"status": job["status"], "filename": job.get("filename"), "error": job.get("error")}


@admin_router.get("/export-pdf/download/{job_id}")
def download_pdf(request, job_id: str):
    _require_approved(request.auth)
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
    if not job:
        raise HttpError(404, "Job_not_found")
    if job["status"] != "done":
        raise HttpError(400, "Job_not_ready")
    filepath = job["file"]
    filename = job["filename"]
    if not filepath or not os.path.exists(filepath):
        raise HttpError(404, "File_not_found")
    with open(filepath, "rb") as f:
        content = f.read()
    try:
        os.unlink(filepath)
    except Exception:
        pass
    with _JOBS_LOCK:
        _JOBS.pop(job_id, None)
    response = HttpResponse(content, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ── Endpoint de Dashboard ─────────────────────────────────────────────────────
@admin_router.get("/dashboard", response={200: dict})
def get_dashboard(
    request,
    id_condominio: Optional[int] = None,
    data_posicao: Optional[str] = None,
    ultimos_5_anos: bool = False,
):
    """
    Retorna dados resumidos para o dashboard.
    ultimos_5_anos=true filtra vencimentos dos últimos 5 anos.
    """
    _require_approved(request.auth)

    from core.superlogica import (
        verificar_condominio,
        buscar_unidades,
        buscar_inadimplentes_condominio,
        _MAX_WORKERS_CONDOMINIOS,
    )
    from django.conf import settings
    from decimal import Decimal
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if not data_posicao:
        data_posicao_fmt   = datetime.today().strftime("%m/%d/%Y")
        data_posicao_label = datetime.today().strftime("%d/%m/%Y")
    else:
        partes = data_posicao.split("/")
        if len(partes) == 3:
            data_posicao_fmt = f"{partes[1]}/{partes[0]}/{partes[2]}"
        else:
            data_posicao_fmt = data_posicao
        data_posicao_label = data_posicao

    # Calcula data_inicio se filtro de 5 anos estiver ativo
    data_inicio = None
    if ultimos_5_anos:
        data_inicio = (datetime.now() - timedelta(days=5 * 365)).strftime("%d/%m/%Y")

    # Cache key inclui o flag de 5 anos
    cache_key = f"{id_condominio or 'todos'}_{data_posicao_label}_{'5a' if ultimos_5_anos else 'all'}"

    with _CACHE_LOCK:
        cached = _DASHBOARD_CACHE.get(cache_key)
        if cached and cached["expires_at"] > datetime.now():
            return 200, cached["data"]

    ids_range    = [id_condominio] if id_condominio else range(1, getattr(settings, "SUPERLOGICA_MAX_ID", 100) + 1)
    total_geral  = Decimal("0")
    condo_valores = {}
    sem_numero   = []
    total_unidades = 0

    def _processar(condo_id):
        acesso, nome = verificar_condominio(condo_id)
        if not acesso:
            return None
        mapa = buscar_unidades(condo_id)
        if not mapa:
            return None
        try:
            resumo, _ = buscar_inadimplentes_condominio(condo_id, data_posicao_fmt, mapa, data_inicio)
        except TypeError:
            # Retrocompatibilidade: superlogica.py ainda sem patch data_inicio
            resumo, _ = buscar_inadimplentes_condominio(condo_id, data_posicao_fmt, mapa)
        if not resumo:
            return None

        condo_total   = Decimal("0")
        sem_num_local = []
        for uid, vals in resumo.items():
            try:
                condo_total += Decimal(str(vals["total"]))
            except Exception:
                pass
            tels     = vals.get("telefones", [])
            t1       = tels[0] if len(tels) > 0 else "s/n"
            t2       = tels[1] if len(tels) > 1 else "s/n"
            dados_uni = mapa.get(uid, {})
            if t1.lower() == "s/n" and t2.lower() == "s/n":
                sem_num_local.append({
                    "condominio": nome or "",
                    "unidade":    dados_uni.get("unidade") or vals.get("nome_pdf", ""),
                    "nome":       dados_uni.get("sacado", ""),
                })

        return nome, float(condo_total.quantize(Decimal("0.01"))), len(resumo), sem_num_local

    with ThreadPoolExecutor(max_workers=_MAX_WORKERS_CONDOMINIOS) as executor:
        futures = {executor.submit(_processar, cid): cid for cid in ids_range}
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    nome_c, val_c, qtd_c, sem_c = result
                    total_geral += Decimal(str(val_c))
                    condo_valores[nome_c] = condo_valores.get(nome_c, 0) + val_c
                    total_unidades += qtd_c
                    sem_numero.extend(sem_c)
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f"Dashboard _processar erro: {e}", exc_info=True)

    maior_condo = max(condo_valores, key=condo_valores.get) if condo_valores else None
    maior_valor = condo_valores.get(maior_condo, 0) if maior_condo else 0

    sem_numero.sort(key=lambda x: (x["condominio"], x["unidade"]))

    condo_ranking = sorted(
        [{"nome": k, "valor": round(v, 2)} for k, v in condo_valores.items()],
        key=lambda x: x["valor"],
        reverse=True,
    )

    resultado = {
        "total_inadimplencia": float(total_geral.quantize(Decimal("0.01"))),
        "total_condominios":   len(condo_valores),
        "total_unidades":      total_unidades,
        "maior_condo_nome":    maior_condo,
        "maior_condo_valor":   round(maior_valor, 2),
        "sem_numero_count":    len(sem_numero),
        "sem_numero":          sem_numero,
        "condo_ranking":       condo_ranking,
        "gerado_em":           datetime.now().strftime("%d/%m/%Y %H:%M"),
        "cache":               False,
    }

    with _CACHE_LOCK:
        _DASHBOARD_CACHE[cache_key] = {
            "data":       {**resultado, "cache": True},
            "expires_at": _proximas_8h(),
        }

    return 200, resultado


# ── Endpoint: listar condomínios ──────────────────────────────────────────────
@admin_router.get("/condominios", response={200: list})
def listar_condominios(request):
    _require_approved(request.auth)

    from django.conf import settings
    import requests as req

    headers = {
        "app_token":    settings.SUPERLOGICA_APP_TOKEN,
        "access_token": settings.SUPERLOGICA_ACCESS_TOKEN,
    }

    max_id     = getattr(settings, "SUPERLOGICA_MAX_ID", 100)
    condominios = []

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _checar(cid):
        try:
            r = req.get(
                f"{settings.SUPERLOGICA_BASE_URL}/unidades",
                headers=headers,
                params={"idCondominio": cid, "pagina": 1, "itensPorPagina": 1},
                timeout=15,
            )
            if r.status_code != 200:
                return None
            dados = r.json()
            if not dados:
                return None
            nome = dados[0].get("st_nome_cond", "").strip()
            if nome:
                return {"id": cid, "nome": nome}
        except Exception:
            pass
        return None

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(_checar, cid): cid for cid in range(1, max_id + 1)}
        for future in as_completed(futures):
            result = future.result()
            if result:
                condominios.append(result)

    condominios.sort(key=lambda x: x["nome"])
    return 200, condominios


# Condomínios cujo jurídico não é da equipe Pratika
JURIDICO_EXTERNO = {18, 7, 19, 9, 25, 16, 20, 65, 27, 32, 22, 45, 46, 47, 3, 48}

@admin_router.get("/sem-numero/xlsx")
def relatorio_sem_numero_xlsx(
    request,
    id_condominio: str = None,
    ultimos_5_anos: bool = False,
    min_inadimplencias: int = 0,
    excluir_juridico_externo: bool = False,
):
    """Gera Excel com unidades sem número cadastrado."""
    _require_admin(request.auth)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        from core.superlogica import _get_headers, verificar_condominio
        from django.conf import settings as _settings
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import requests as _req

        def _checar(cid):
            try:
                headers = _get_headers()
                r = _req.get(f"{_settings.SUPERLOGICA_BASE_URL}/unidades", headers=headers,
                             params={"idCondominio": cid, "pagina": 1, "itensPorPagina": 1}, timeout=10)
                if r.status_code != 200 or not r.json():
                    return None
                nome = r.json()[0].get("st_nome_cond", "").strip()
                return {"id": cid, "nome": nome} if nome else None
            except Exception:
                return None

        max_id = getattr(_settings, "SUPERLOGICA_MAX_ID", 100)
        if id_condominio:
            ids_lista = [int(p.strip()) for p in id_condominio.split(",") if p.strip()]
            condominios = []
            for cid in ids_lista:
                acesso, nome = verificar_condominio(cid)
                if acesso:
                    condominios.append({"id": cid, "nome": nome})
        else:
            condominios = []
            with ThreadPoolExecutor(max_workers=10) as ex:
                futs = {ex.submit(_checar, cid): cid for cid in range(1, max_id + 1)}
                for f in as_completed(futs):
                    r = f.result()
                    if r:
                        condominios.append(r)
            condominios.sort(key=lambda x: x["nome"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Sem Número"

        headers = ["Condomínio", "Bloco", "Unidade", "Nome", "Qtd Inadimpl."]
        header_fill = PatternFill(start_color="006837", end_color="006837", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        from datetime import datetime as _dt
        from core.superlogica import buscar_inadimplentes_condominio
        data_hoje = _dt.today().strftime("%m/%d/%Y")
        if excluir_juridico_externo:
            condominios = [c for c in condominios if int(c["id"]) not in JURIDICO_EXTERNO]

        from datetime import datetime as _dt2, timedelta as _td
        data_inicio_filtro = None
        if ultimos_5_anos:
            data_inicio_filtro = (_dt2.now() - _td(days=5*365)).date()

        row_idx = 2
        for condo in condominios:
            mapa = buscar_unidades(condo["id"])
            try:
                resumo_inad, det_inad = buscar_inadimplentes_condominio(condo["id"], data_hoje, mapa)
                contagem = {}
                for row in (det_inad or []):
                    k = str(row.get("id_unidade") or "")
                    if k:
                        contagem[k] = contagem.get(k, 0) + 1
            except Exception:
                resumo_inad = {}
                contagem = {}
            for uid, dados in mapa.items():
                if not dados.get("telefones"):
                    qtd = contagem.get(str(uid), 0)
                    if min_inadimplencias > 0 and qtd < min_inadimplencias:
                        continue
                    for col, val in enumerate([condo["nome"], dados.get("bloco",""), dados.get("unidade",""), dados.get("sacado",""), qtd], 1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                    row_idx += 1

        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 15

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="sem_numero.xlsx"'
        return response
    except Exception as e:
        raise HttpError(500, str(e))


@admin_router.get("/sem-numero/pdf")
def relatorio_sem_numero_pdf(
    request,
    id_condominio: str = None,
    ultimos_5_anos: bool = False,
    min_inadimplencias: int = 0,
    excluir_juridico_externo: bool = False,
):
    """Gera PDF com unidades sem número cadastrado."""
    _require_admin(request.auth)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        import os

        COR_VERDE  = colors.HexColor("#006837")
        COR_BRANCO = colors.white
        COR_ZEBRA  = colors.HexColor("#F2F2F2")
        COR_TEXTO  = colors.HexColor("#1A1A1A")
        COR_CINZA  = colors.HexColor("#666666")

        style_cell      = ParagraphStyle("cell", fontSize=8, textColor=COR_TEXTO, fontName="Helvetica", leading=10)
        style_cell_bold = ParagraphStyle("bold", fontSize=8, textColor=COR_BRANCO, fontName="Helvetica-Bold", leading=10, alignment=TA_CENTER)
        style_title     = ParagraphStyle("titulo", fontSize=16, textColor=COR_VERDE, fontName="Helvetica-Bold", spaceAfter=4)
        style_sub       = ParagraphStyle("sub", fontSize=9, textColor=COR_CINZA, fontName="Helvetica", spaceAfter=2)
        style_section   = ParagraphStyle("section", fontSize=10, textColor=COR_VERDE, fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=10)

        def p(txt, style=None):
            return Paragraph(str(txt) if txt else "", style or style_cell)

        from core.superlogica import _get_headers, verificar_condominio
        from django.conf import settings as _settings
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import requests as _req

        def _checar(cid):
            try:
                headers = _get_headers()
                r = _req.get(f"{_settings.SUPERLOGICA_BASE_URL}/unidades", headers=headers,
                             params={"idCondominio": cid, "pagina": 1, "itensPorPagina": 1}, timeout=10)
                if r.status_code != 200 or not r.json():
                    return None
                nome = r.json()[0].get("st_nome_cond", "").strip()
                return {"id": cid, "nome": nome} if nome else None
            except Exception:
                return None

        max_id = getattr(_settings, "SUPERLOGICA_MAX_ID", 100)
        if id_condominio:
            ids_lista = [int(p.strip()) for p in id_condominio.split(",") if p.strip()]
            condominios = []
            for cid in ids_lista:
                acesso, nome = verificar_condominio(cid)
                if acesso:
                    condominios.append({"id": cid, "nome": nome})
        else:
            condominios = []
            with ThreadPoolExecutor(max_workers=10) as ex:
                futs = {ex.submit(_checar, cid): cid for cid in range(1, max_id + 1)}
                for f in as_completed(futs):
                    r = f.result()
                    if r:
                        condominios.append(r)
            condominios.sort(key=lambda x: x["nome"])

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
        story = []

        _logo_path = os.path.join(os.path.dirname(__file__), "logo_pratika.png")
        if os.path.exists(_logo_path):
            from reportlab.platypus import Image as RLImage
            logo_img = RLImage(_logo_path, width=3.5*cm, height=1.8*cm)
        else:
            logo_img = Paragraph("", style_sub)

        header_table = Table([[logo_img, Paragraph("Unidades sem Número Cadastrado", style_title)]], colWidths=[4*cm, None])
        header_table.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("LEFTPADDING", (0,0), (-1,-1), 0)]))
        story.append(header_table)

        from datetime import datetime
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_sub))
        story.append(HRFlowable(width="100%", thickness=2, color=COR_VERDE, spaceAfter=10))

        if excluir_juridico_externo:
            condominios = [c for c in condominios if int(c["id"]) not in JURIDICO_EXTERNO]

        total = 0
        for condo in condominios:
            mapa = buscar_unidades(condo["id"])
            linhas = [(dados.get("bloco",""), dados.get("unidade",""), dados.get("sacado",""), uid) for uid, dados in mapa.items() if not dados.get("telefones")]
            if not linhas:
                continue
            total += len(linhas)
            story.append(Paragraph(f"{condo['nome']}", style_section))
            from datetime import datetime as _dt
            from core.superlogica import buscar_inadimplentes_condominio
            data_hoje = _dt.today().strftime("%m/%d/%Y")
            try:
                resumo_inad, det_inad = buscar_inadimplentes_condominio(condo["id"], data_hoje, mapa)
                contagem = {}
                for row in (det_inad or []):
                    k = str(row.get("id_unidade") or "")
                    if k:
                        contagem[k] = contagem.get(k, 0) + 1
            except Exception:
                contagem = {}
            cab = [p("Bloco", style_cell_bold), p("Unidade", style_cell_bold), p("Nome", style_cell_bold), p("Qtd Inadimpl.", style_cell_bold)]
            dados_tab = [cab]
            for i, (bloco, uni, nome, uid) in enumerate(sorted(linhas)):
                qtd = contagem.get(str(uid), 0)
                if min_inadimplencias > 0 and qtd < min_inadimplencias:
                    continue
                dados_tab.append([p(bloco), p(uni), p(nome), p(str(qtd) if qtd else "-")])
            if len(dados_tab) <= 1:
                continue
            n_rows = len(dados_tab) - 1
            zebra = [("BACKGROUND", (0, idx+1), (-1, idx+1), COR_BRANCO if idx%2==0 else COR_ZEBRA) for idx in range(n_rows)]
            t = Table(dados_tab, colWidths=[3*cm, 3*cm, 9*cm, 2*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), COR_VERDE),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ] + zebra))
            story.append(t)
            story.append(Spacer(1, 0.3*cm))

        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="sem_numero.pdf"'
        return response
    except Exception as e:
        raise HttpError(500, str(e))
