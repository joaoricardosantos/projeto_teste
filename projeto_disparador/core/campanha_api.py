"""
Endpoints para gerenciamento de campanhas e reenvio de mensagens.
"""
import io
import logging
from typing import List, Optional
from ninja import Router, Schema
from ninja.errors import HttpError
from pydantic import UUID4
from django.http import HttpResponse
from core.auth import JWTAuth

logger = logging.getLogger(__name__)
campanha_router = Router(auth=JWTAuth())


def _fmt(dt):
    """Formata datetime para fuso America/Sao_Paulo — só a data, sem horário."""
    if not dt:
        return None
    from django.utils import timezone as tz
    import zoneinfo
    sp = zoneinfo.ZoneInfo("America/Sao_Paulo")
    local = dt.astimezone(sp)
    return local.strftime("%d/%m/%Y")


def _aggregate_bulk_results(results: list) -> dict:
    success  = sum(1 for r in results if r.get("status") == "success")
    errors   = sum(1 for r in results if r.get("status") != "success")
    failures = [
        {"phone": r.get("phone", ""), "error": r.get("error", "Erro desconhecido")}
        for r in results
        if r.get("status") != "success"
    ]
    return {"success": success, "errors": errors, "failures": failures}


@campanha_router.get("/", response={200: list})
def listar_campanhas(request):
    from core.models import Campanha
    campanhas = Campanha.objects.all()
    resultado = []
    for c in campanhas:
        respondidos = c.mensagens.filter(status="respondido").count()
        enviados    = c.mensagens.filter(status="enviado").count()
        erros       = c.mensagens.filter(status="erro").count()
        resultado.append({
            "id":               str(c.id),
            "nome":             c.nome,
            "criada_em":        _fmt(c.criada_em),
            "total_enviados":   c.total_enviados,
            "total_erros":      c.total_erros,
            "total_sem_numero": c.total_sem_numero,
            "respondidos":      respondidos,
            "aguardando":       enviados,
            "erros_envio":      erros,
        })
    return 200, resultado


@campanha_router.get("/{campanha_id}/mensagens", response={200: list})
def listar_mensagens(request, campanha_id: UUID4, status: Optional[str] = None):
    from core.models import Campanha, MensagemEnviada
    try:
        campanha = Campanha.objects.get(id=campanha_id)
    except Campanha.DoesNotExist:
        raise HttpError(404, "Campanha_not_found")

    qs = campanha.mensagens.all()
    if status:
        qs = qs.filter(status=status)

    return 200, [
        {
            "id":           str(m.id),
            "condominio":   m.condominio,
            "unidade":      m.unidade,
            "nome":         m.nome,
            "telefone":     m.telefone,
            "status":       m.status,
            "enviada_em":   _fmt(m.enviada_em),
            "respondida_em":_fmt(m.respondida_em),
            "resposta":     m.resposta,
        }
        for m in qs
    ]


# ── Renomear campanha ─────────────────────────────────────────────────────────

class RenomearIn(Schema):
    nome: str


@campanha_router.patch("/{campanha_id}/renomear", response={200: dict})
def renomear_campanha(request, campanha_id: UUID4, payload: RenomearIn):
    """Renomeia uma campanha existente."""
    from core.models import Campanha
    nome = payload.nome.strip()
    if not nome:
        raise HttpError(400, "Nome_nao_pode_ser_vazio")
    try:
        campanha = Campanha.objects.get(id=campanha_id)
    except Campanha.DoesNotExist:
        raise HttpError(404, "Campanha_not_found")
    campanha.nome = nome
    campanha.save(update_fields=["nome"])
    return 200, {"id": str(campanha.id), "nome": campanha.nome}


# ── Deletar campanha ──────────────────────────────────────────────────────────

@campanha_router.delete("/{campanha_id}", response={200: dict})
def deletar_campanha(request, campanha_id: UUID4):
    if not request.auth.is_staff and not request.auth.is_superuser:
        raise HttpError(403, "Admin_privileges_required")
    from core.models import Campanha
    try:
        campanha = Campanha.objects.get(id=campanha_id)
        campanha.delete()
        return 200, {"message": "Campanha removida com sucesso"}
    except Campanha.DoesNotExist:
        raise HttpError(404, "Campanha_not_found")


# ── Reenviar mensagens ────────────────────────────────────────────────────────

class ReenviarIn(Schema):
    ids: List[str]
    template_id: Optional[str] = None


@campanha_router.post("/{campanha_id}/reenviar", response={200: dict})
def reenviar_mensagens(request, campanha_id: UUID4, payload: ReenviarIn):
    from core.models import Campanha, MensagemEnviada, MessageTemplate
    from core.evolution_service import send_whatsapp_bulk
    from core.services import _render_template

    try:
        campanha = Campanha.objects.get(id=campanha_id)
    except Campanha.DoesNotExist:
        raise HttpError(404, "Campanha_not_found")

    ids = payload.ids
    if not ids:
        raise HttpError(400, "Nenhum_ID_informado")

    mensagens = MensagemEnviada.objects.filter(id__in=ids, campanha=campanha)
    if not mensagens.exists():
        raise HttpError(404, "Mensagens_not_found")

    template_body = None
    if payload.template_id:
        try:
            t = MessageTemplate.objects.get(id=payload.template_id)
            template_body = t.body
        except MessageTemplate.DoesNotExist:
            raise HttpError(404, "Template_not_found")

    contacts = []
    for m in mensagens:
        msg = _render_template(template_body, condo_name=m.condominio, unidade=m.unidade, nome=m.nome) \
              if template_body else m.mensagem
        contacts.append({"phone": m.telefone, "message": msg, "_id": str(m.id)})

    result = _aggregate_bulk_results(send_whatsapp_bulk(contacts))

    from django.utils import timezone
    MensagemEnviada.objects.filter(id__in=ids).update(
        status=MensagemEnviada.STATUS_ENVIADO,
        enviada_em=timezone.now(),
        respondida_em=None,
        resposta="",
    )

    return 200, {"success": result["success"], "errors": result["errors"], "failures": result["failures"]}


# ── Relatório de disparo ───────────────────────────────────────────────────────

def _nome_arquivo_campanha(campanha):
    from django.db.models import Count
    import zoneinfo
    condo_top = (
        campanha.mensagens.values("condominio")
        .annotate(total=Count("id"))
        .order_by("-total")
        .first()
    )
    condo_nome = (condo_top["condominio"] if condo_top else campanha.nome) or campanha.nome
    data_disparo = campanha.criada_em.astimezone(
        zoneinfo.ZoneInfo("America/Sao_Paulo")
    ).strftime("%d-%m-%Y")
    return f"{condo_nome}_{data_disparo}".replace(" ", "_").replace("/", "-")[:80]


@campanha_router.get("/{campanha_id}/relatorio-excel", auth=JWTAuth())
def relatorio_excel(request, campanha_id: UUID4):
    """Exporta relatório Excel de uma campanha: respondidos, aguardando e sem número."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from core.models import Campanha
    from core.superlogica import _estilizar_cabecalho, _ajustar_larguras, _aplicar_grade_e_zebra

    try:
        campanha = Campanha.objects.get(id=campanha_id)
    except Campanha.DoesNotExist:
        raise HttpError(404, "Campanha_not_found")

    respondidos = list(campanha.mensagens.filter(status="respondido").order_by("condominio", "nome"))
    aguardando  = list(campanha.mensagens.filter(status="enviado").order_by("condominio", "nome"))
    sem_numero  = list(campanha.mensagens.filter(status="erro").order_by("condominio", "nome"))

    VERDE   = "006837"
    FILL_PAR   = PatternFill("solid", fgColor="EAF4EE")
    FILL_IMPAR = PatternFill("solid", fgColor="FFFFFF")
    borda = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    HEADERS     = ["Condomínio", "Unidade", "Nome", "Telefone", "Enviada em", "Resposta / Motivo"]
    WIDTHS      = [35, 10, 30, 18, 14, 55]

    def _make_sheet(ws, rows, status_label, sem_numero=False):
        ws.append(HEADERS)
        for i, m in enumerate(rows):
            ws.append([
                m.condominio,
                m.unidade,
                m.nome,
                m.telefone or "—",
                _fmt(m.enviada_em) or "—",
                m.mensagem if sem_numero else (m.resposta or "—"),
            ])
            fill = FILL_PAR if i % 2 == 0 else FILL_IMPAR
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=i + 2, column=col)
                cell.fill      = fill
                cell.border    = borda
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
                cell.font      = Font(size=9)
            ws.row_dimensions[i + 2].height = 30

        _estilizar_cabecalho(ws)
        for col, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    wb = Workbook()

    # Aba Resumo (índice 0)
    ws_res = wb.active
    ws_res.title = "Resumo"
    ws_res.append(["Status", "Quantidade"])
    _estilizar_cabecalho(ws_res)

    total = len(respondidos) + len(aguardando) + len(sem_numero)
    resumo_data = [
        ("Responderam", len(respondidos), "4CAF50"),
        ("Aguardando",  len(aguardando),  "FFC107"),
        ("Sem número",  len(sem_numero),  "F44336"),
        ("Total",       total,            VERDE),
    ]
    for i, (label, valor, cor) in enumerate(resumo_data):
        r = i + 2
        c1 = ws_res.cell(row=r, column=1, value=label)
        c1.font      = Font(bold=True, size=11, color="FFFFFF")
        c1.fill      = PatternFill("solid", fgColor=cor)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border    = borda
        c2 = ws_res.cell(row=r, column=2, value=valor)
        c2.font      = Font(bold=True, size=14, color=cor)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border    = borda
        ws_res.row_dimensions[r].height = 32

    ws_res.column_dimensions["A"].width = 20
    ws_res.column_dimensions["B"].width = 14

    # Abas de dados
    ws_resp = wb.create_sheet("Respondidos")
    ws_ag   = wb.create_sheet("Aguardando")
    ws_sn   = wb.create_sheet("Sem número")

    _make_sheet(ws_resp, respondidos, "Respondido")
    _make_sheet(ws_ag,   aguardando,  "Aguardando")
    _make_sheet(ws_sn,   sem_numero,  "Sem número", sem_numero=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = _nome_arquivo_campanha(campanha)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="relatorio_{nome_arquivo}.xlsx"'
    return response


@campanha_router.get("/{campanha_id}/relatorio-pdf", auth=JWTAuth())
def relatorio_pdf(request, campanha_id: UUID4):
    """Exporta relatório PDF de uma campanha: respondidos, aguardando e sem número."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HttpError(500, "reportlab não instalado")

    from core.models import Campanha

    try:
        campanha = Campanha.objects.get(id=campanha_id)
    except Campanha.DoesNotExist:
        raise HttpError(404, "Campanha_not_found")

    respondidos = list(campanha.mensagens.filter(status="respondido").order_by("condominio", "nome"))
    aguardando  = list(campanha.mensagens.filter(status="enviado").order_by("condominio", "nome"))
    sem_numero  = list(campanha.mensagens.filter(status="erro").order_by("condominio", "nome"))

    VERDE_ESCURO = colors.HexColor("#006837")
    VERDE_CLARO  = colors.HexColor("#EAF4EE")
    BRANCO       = colors.white
    CINZA        = colors.HexColor("#F5F5F5")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=14, textColor=BRANCO, fontName="Helvetica-Bold", spaceAfter=4)
    sec_style   = ParagraphStyle("sec",   fontSize=11, textColor=BRANCO, fontName="Helvetica-Bold", spaceAfter=2)
    small       = ParagraphStyle("small", fontSize=7, leading=9)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    def _section_title(text, color=VERDE_ESCURO):
        data = [[Paragraph(text, sec_style)]]
        t = Table(data, colWidths=[doc.width])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), color),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))
        return t

    def _make_table(rows, status_label, sem_numero=False):
        if not rows:
            return Paragraph("<i>Nenhum registro.</i>", styles["Normal"])
        header = ["Condomínio", "Unidade", "Nome", "Telefone", "Enviada em", "Resposta / Motivo"]
        col_w  = [6.5*cm, 2*cm, 5.5*cm, 3*cm, 2.5*cm, 8.5*cm]
        data   = [header]
        for m in rows:
            ultima_col = m.mensagem if sem_numero else (m.resposta or "—")
            data.append([
                Paragraph(m.condominio or "—", small),
                m.unidade or "—",
                Paragraph(m.nome or "—", small),
                m.telefone or "—",
                _fmt(m.enviada_em) or "—",
                Paragraph(ultima_col or "—", small),
            ])
        t = Table(data, colWidths=col_w, repeatRows=1)
        style = TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  VERDE_ESCURO),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  BRANCO),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  8),
            ("FONTSIZE",      (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [VERDE_CLARO, BRANCO]),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
        t.setStyle(style)
        return t

    # Título principal
    story.append(_section_title(f"Relatório de Disparo — {campanha.nome}", VERDE_ESCURO))
    story.append(Spacer(1, 0.3*cm))

    # Resumo
    resumo_data = [
        ["Status", "Quantidade"],
        ["Responderam", str(len(respondidos))],
        ["Aguardando",  str(len(aguardando))],
        ["Sem número",  str(len(sem_numero))],
        ["Total",       str(len(respondidos) + len(aguardando) + len(sem_numero))],
    ]
    t_resumo = Table(resumo_data, colWidths=[5*cm, 3*cm])
    t_resumo.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  VERDE_ESCURO),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  BRANCO),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [VERDE_CLARO, BRANCO]),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN",         (1, 0), (1, -1),  "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t_resumo)
    story.append(Spacer(1, 0.5*cm))

    # Seções
    for titulo, rows, label, is_sem_num in [
        ("Respondidos", respondidos, "Respondido", False),
        ("Aguardando",  aguardando,  "Aguardando", False),
        ("Sem número",  sem_numero,  "Sem número", True),
    ]:
        story.append(_section_title(f"{titulo} ({len(rows)})"))
        story.append(Spacer(1, 0.2*cm))
        story.append(_make_table(rows, label, sem_numero=is_sem_num))
        story.append(Spacer(1, 0.5*cm))

    doc.build(story)
    buf.seek(0)

    nome_arquivo = _nome_arquivo_campanha(campanha)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="relatorio_{nome_arquivo}.pdf"'
    return response