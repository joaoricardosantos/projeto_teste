import csv
import openpyxl
from io import StringIO, BytesIO
from django.conf import settings
from core.evolution_service import send_whatsapp_bulk


def _render_template(
    body: str,
    condo_name: str = "",
    unidade: str = "",
    nome: str = "",
    qtd_inadimpl: str = "",
    competencia: str = "",
    vencimento: str = "",
    valor: str = "",
) -> str:
    return (
        body
        .replace("{{condominio}}",  condo_name)
        .replace("{{unidade}}",     unidade)
        .replace("{{nome}}",        nome)
        .replace("{{qtd}}",         str(qtd_inadimpl))
        .replace("{{competencia}}", competencia)
        .replace("{{vencimento}}",  vencimento)
        .replace("{{valor}}",       valor)
    )


def _build_default_message(
    condo_name: str,
    unidade: str,
    nome: str,
    vencimento: str,
    competencia: str,
    valor: str,
    qtd: str,
) -> str:
    return (
        f"Olá{', ' + nome if nome else ''}! Identificamos débito em aberto referente à unidade "
        f"*{unidade}* do condomínio *{condo_name}*.\n"
        f"Competência: *{competencia}* | Vencimento: *{vencimento}*\n"
        f"Qtd de inadimplências: *{qtd}*\n"
        f"Valor total com encargos: *{valor}*.\n"
        f"Entre em contato para regularização."
    )


def _sem_numero(tel: str) -> bool:
    """Retorna True se o valor da célula indica ausência de número."""
    return not tel or tel.strip().lower() in ("s/n", "s/ n", "", "none", "nan", "-")


def _agregar_resultado(raw, extra_errors: int = 0, failures_no_phone: list = None) -> dict:
    """
    send_whatsapp_bulk pode retornar list OU dict dependendo da versão.
    Normaliza sempre para: {"success": int, "errors": int, "failures": list, "sem_numero": list}
    """
    if failures_no_phone is None:
        failures_no_phone = []

    if isinstance(raw, list):
        success  = sum(1 for r in raw if r.get("status") == "success")
        failures = [
            {"phone": r.get("phone", ""), "error": r.get("error", "Erro desconhecido")}
            for r in raw if r.get("status") != "success"
        ]
        errors = len(failures)
        result = {"success": success, "errors": errors + extra_errors, "failures": failures}
    elif isinstance(raw, dict):
        result = raw
        result["errors"] = result.get("errors", 0) + extra_errors
        result.setdefault("failures", [])
    else:
        result = {"success": 0, "errors": extra_errors, "failures": []}

    result["sem_numero"] = failures_no_phone
    result["failures"] = result.get("failures", []) + [
        {"phone": "—", "error": f"{f['unidade']} ({f['nome']}): {f['motivo']}"}
        for f in failures_no_phone
    ]
    return result


def _registrar_campanha(contacts: list, result: dict, failures_no_phone: list):
    """Registra campanha e mensagens enviadas no banco de dados."""
    try:
        from core.models import Campanha, MensagemEnviada
        from django.utils import timezone as tz
        import re as _re

        campanha_nome = "Novo disparo"
        campanha = Campanha.objects.create(
            nome=campanha_nome,
            total_enviados=result.get("success", 0),
            total_erros=result.get("errors", 0),
            total_sem_numero=len(failures_no_phone),
        )
        result["campanha_id"] = str(campanha.id)
        result["campanha_nome"] = campanha_nome

        def _norm(p): return _re.sub(r"\D", "", p)
        objs = [
            MensagemEnviada(
                campanha=campanha,
                condominio=c.get("condominio", ""),
                unidade=c.get("unidade", ""),
                nome=c.get("nome", ""),
                telefone=_norm(c.get("phone", "")),
                mensagem=c.get("message", ""),
                status=MensagemEnviada.STATUS_ENVIADO,
            )
            for c in contacts
        ]
        MensagemEnviada.objects.bulk_create(objs)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Erro ao salvar campanha: {e}")


def _ler_xlsx_como_relatorio(conteudo: bytes, failures_no_phone: list) -> tuple:
    """
    Lê XLSX no formato relatório do sistema (aba Resumo).
    Popula failures_no_phone com unidades sem número.
    Retorna (lista de contatos prontos para envio, contagem de erros).
    """
    wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb["Resumo"] if "Resumo" in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], 0

    headers_lower = [str(c).strip().lower() if c is not None else "" for c in header_row]

    def _idx(*names):
        for n in names:
            for i, h in enumerate(headers_lower):
                if h == n.lower():
                    return i
        return None

    idx_condo   = _idx("condomínio", "condominio")
    idx_unidade = _idx("unidade")
    idx_nome    = _idx("nome")
    idx_tel1    = _idx("telefone 1", "telefone1")
    idx_tel2    = _idx("telefone 2", "telefone2")
    idx_total   = _idx("total")
    idx_qtd     = _idx("qtd inadimpl.", "qtd inadimpl")
    idx_venc    = _idx("vencimento")
    idx_comp    = _idx("competência", "competencia")

    def _cell(row, idx):
        if idx is not None and idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip()
        return ""

    linhas = []
    erros = 0

    for row in rows_iter:
        if not any(row):
            continue
        first = _cell(row, 0).upper()
        if first in ("TOTAL GERAL", "TOTAL"):
            continue

        condo_name  = _cell(row, idx_condo)
        unidade     = _cell(row, idx_unidade)
        nome        = _cell(row, idx_nome)
        qtd         = _cell(row, idx_qtd)
        vencimento  = _cell(row, idx_venc)
        competencia = _cell(row, idx_comp)

        tel1 = _cell(row, idx_tel1)
        tel2 = _cell(row, idx_tel2)

        total_raw = row[idx_total] if idx_total is not None and idx_total < len(row) else None
        try:
            valor_fmt = f"R$ {float(total_raw):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            valor_fmt = str(total_raw) if total_raw else ""

        # Escolhe telefone disponível (ignora s/n)
        telefone = None
        if not _sem_numero(tel1):
            telefone = tel1
        elif not _sem_numero(tel2):
            telefone = tel2

        if not telefone:
            # ── Sem número: registra em failures_no_phone para exibição detalhada ──
            failures_no_phone.append({
                "unidade": unidade,
                "nome":    nome,
                "motivo":  "Sem número cadastrado",
            })
            erros += 1
            continue

        linhas.append({
            "condominio":  condo_name,
            "contato":     telefone,
            "valor_debito": valor_fmt,
            "unidade":     unidade,
            "nome":        nome,
            "qtd":         qtd,
            "vencimento":  vencimento,
            "competencia": competencia,
        })

    wb.close()
    return linhas, erros


def _ler_linhas_legado(file_obj, failures_no_phone: list) -> tuple:
    """
    Lê CSV ou XLSX (formato legado ou relatório).
    Popula failures_no_phone com unidades sem número quando aplicável.
    Retorna (lista de dicts de contatos, contagem de erros).
    """
    nome_arquivo = getattr(file_obj, "name", "") or ""
    is_xlsx = nome_arquivo.lower().endswith(".xlsx")

    if is_xlsx:
        conteudo = file_obj.read()

        # Detecta formato pelo cabeçalho
        wb_check = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        ws_check = wb_check["Resumo"] if "Resumo" in wb_check.sheetnames else wb_check.active
        header_check = []
        for row in ws_check.iter_rows(min_row=1, max_row=1, values_only=True):
            header_check = [str(c).strip().lower() if c else "" for c in row]
            break
        wb_check.close()

        is_relatorio = any(h in header_check for h in ["telefone 1", "telefone1", "unidade", "qtd inadimpl."])

        if is_relatorio:
            return _ler_xlsx_como_relatorio(conteudo, failures_no_phone)

        # Formato legado XLSX
        wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return [], 0

        headers_lower = [str(c).strip().lower() if c is not None else "" for c in header_row]

        def _idx_leg(*names):
            for n in names:
                if n in headers_lower:
                    return headers_lower.index(n)
            return None

        idx_condo   = _idx_leg("condominio")
        idx_contato = _idx_leg("contato")
        idx_valor   = _idx_leg("valor_debito")

        linhas, erros = [], 0
        for row in rows_iter:
            if not any(row):
                continue

            def _cell(idx):
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return ""

            condo   = _cell(idx_condo)
            contato = _cell(idx_contato)
            valor   = _cell(idx_valor)

            if condo and contato and valor:
                linhas.append({"condominio": condo, "contato": contato, "valor_debito": valor})
            else:
                erros += 1
        wb.close()
        return linhas, erros

    else:
        # CSV
        conteudo = file_obj.read()
        amostra = conteudo[:2048].decode("utf-8", errors="replace")
        sep = ";" if amostra.count(";") > amostra.count(",") else ","
        reader = csv.DictReader(StringIO(conteudo.decode("utf-8")), delimiter=sep)
        linhas, erros = [], 0
        for row in reader:
            condo   = (row.get("condominio") or "").strip()
            contato = (row.get("contato") or "").strip()
            valor   = (row.get("valor_debito") or "").strip()
            if condo and contato and valor:
                linhas.append({"condominio": condo, "contato": contato, "valor_debito": valor})
            else:
                erros += 1
        return linhas, erros


def process_defaulters_spreadsheet(file_obj):
    """
    Processa CSV ou XLSX com mensagem padrão e envia via WhatsApp.
    Aceita formato legado (condominio/contato/valor_debito)
    e formato relatório gerado pelo sistema (aba Resumo com Telefone 1/2).
    """
    failures_no_phone = []
    linhas, error_count = _ler_linhas_legado(file_obj, failures_no_phone)

    contacts = []
    for l in linhas:
        if l.get("unidade"):
            message = _build_default_message(
                l["condominio"], l["unidade"], l.get("nome", ""),
                l.get("vencimento", ""), l.get("competencia", ""),
                l["valor_debito"], l.get("qtd", ""),
            )
        else:
            message = _build_default_message(l["condominio"], "", "", "", "", l["valor_debito"], "")
        contacts.append({
            "phone": l["contato"], "message": message,
            "condominio": l["condominio"], "unidade": l.get("unidade", ""), "nome": l.get("nome", ""),
        })

    raw = send_whatsapp_bulk(contacts)
    result = _agregar_resultado(raw, extra_errors=error_count, failures_no_phone=failures_no_phone)
    _registrar_campanha(contacts, result, failures_no_phone)
    return result


def process_defaulters_with_template(file_obj, template_id: str):
    """
    Processa CSV ou XLSX com template específico e envia via WhatsApp.
    Aceita formato legado e formato relatório.
    """
    from core.models import MessageTemplate

    try:
        template = MessageTemplate.objects.get(id=template_id)
    except MessageTemplate.DoesNotExist:
        raise ValueError("Template_not_found")

    failures_no_phone = []
    linhas, error_count = _ler_linhas_legado(file_obj, failures_no_phone)

    contacts = []
    for l in linhas:
        message = _render_template(
            template.body,
            condo_name=l["condominio"],
            unidade=l.get("unidade", ""),
            nome=l.get("nome", l["contato"]),
            qtd_inadimpl=l.get("qtd", ""),
            competencia=l.get("competencia", ""),
            vencimento=l.get("vencimento", ""),
            valor=l["valor_debito"],
        )
        contacts.append({
            "phone": l["contato"], "message": message,
            "condominio": l["condominio"], "unidade": l.get("unidade", ""), "nome": l.get("nome", ""),
        })

    raw = send_whatsapp_bulk(contacts)
    result = _agregar_resultado(raw, extra_errors=error_count, failures_no_phone=failures_no_phone)
    _registrar_campanha(contacts, result, failures_no_phone)
    return result


def process_excel_report_dispatch(file_obj, template_id: str = None):
    """
    Lê o Excel gerado pelo relatório (aba Resumo) e envia WhatsApp
    para Telefone 1 e Telefone 2 de cada unidade inadimplente.
    """
    from core.models import MessageTemplate

    template_body = None
    if template_id:
        try:
            t = MessageTemplate.objects.get(id=template_id)
            template_body = t.body
        except MessageTemplate.DoesNotExist:
            raise ValueError("Template_not_found")

    wb = openpyxl.load_workbook(file_obj)
    ws = wb["Resumo"] if "Resumo" in wb.sheetnames else wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_condo        = col("Condomínio")
    idx_unidade      = col("Unidade")
    idx_nome         = col("Nome")
    idx_tel1         = col("Telefone 1")
    idx_tel2         = col("Telefone 2")
    idx_qtd          = col("Qtd Inadimpl.")
    idx_vencimento   = col("Vencimento")
    idx_competencia  = col("Competência")
    idx_total        = col("Total")
    idx_telefones_legado = col("Telefones")

    if idx_tel1 is None and idx_telefones_legado is None:
        raise ValueError("Coluna_Telefone_nao_encontrada")

    contacts = []
    error_count = 0
    failures_no_phone = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_val = str(row[0]).strip() if row[0] else ""
        if first_val.upper() in ("TOTAL GERAL", "TOTAL"):
            continue

        condo_name  = str(row[idx_condo])       if idx_condo      is not None and row[idx_condo]      else ""
        unidade     = str(row[idx_unidade])     if idx_unidade    is not None and row[idx_unidade]    else ""
        nome        = str(row[idx_nome])        if idx_nome       is not None and row[idx_nome]       else ""
        qtd         = str(row[idx_qtd])         if idx_qtd        is not None and row[idx_qtd]        else ""
        vencimento  = str(row[idx_vencimento])  if idx_vencimento is not None and row[idx_vencimento] else ""
        competencia = str(row[idx_competencia]) if idx_competencia is not None and row[idx_competencia] else ""
        total_raw   = row[idx_total]            if idx_total      is not None else None

        try:
            valor_fmt = f"R$ {float(total_raw):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            valor_fmt = str(total_raw) if total_raw else ""

        if template_body:
            message = _render_template(
                template_body, condo_name=condo_name, unidade=unidade, nome=nome,
                qtd_inadimpl=qtd, competencia=competencia, vencimento=vencimento, valor=valor_fmt,
            )
        else:
            message = _build_default_message(condo_name, unidade, nome, vencimento, competencia, valor_fmt, qtd)

        if idx_tel1 is not None:
            t1 = str(row[idx_tel1]).strip() if row[idx_tel1] else ""
            t2 = str(row[idx_tel2]).strip() if idx_tel2 is not None and row[idx_tel2] else ""

            if not _sem_numero(t1):
                contacts.append({"phone": t1, "message": message,
                    "condominio": condo_name, "unidade": unidade, "nome": nome})
            elif not _sem_numero(t2):
                contacts.append({"phone": t2, "message": message,
                    "condominio": condo_name, "unidade": unidade, "nome": nome})
            else:
                error_count += 1
                failures_no_phone.append({
                    "unidade": unidade,
                    "nome":    nome,
                    "motivo":  "Sem número cadastrado",
                })

        elif idx_telefones_legado is not None:
            raw_tel = row[idx_telefones_legado]
            if raw_tel:
                for phone in [t.strip() for t in str(raw_tel).split("|") if t.strip()]:
                    contacts.append({"phone": phone, "message": message,
                        "condominio": condo_name, "unidade": unidade, "nome": nome})
            else:
                error_count += 1
                failures_no_phone.append({
                    "unidade": unidade,
                    "nome":    nome,
                    "motivo":  "Sem número cadastrado",
                })
        else:
            error_count += 1

    raw = send_whatsapp_bulk(contacts)
    result = _agregar_resultado(raw, extra_errors=error_count, failures_no_phone=failures_no_phone)
    _registrar_campanha(contacts, result, failures_no_phone)
    return result