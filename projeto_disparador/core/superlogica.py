import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from io import BytesIO
from typing import Optional

import requests
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _get_headers() -> dict:
    return {
        "Content-Type": "application/json",
        "app_token": settings.SUPERLOGICA_APP_TOKEN,
        "access_token": settings.SUPERLOGICA_ACCESS_TOKEN,
    }


def verificar_condominio(id_condominio: int):
    response = requests.get(
        f"{settings.SUPERLOGICA_BASE_URL}/unidades",
        headers=_get_headers(),
        params={"idCondominio": id_condominio, "pagina": 1, "itensPorPagina": 1},
        timeout=20,
    )
    if response.status_code != 200:
        return False, None
    dados = response.json()
    nome_condominio = None
    if dados and isinstance(dados, list):
        nome_condominio = dados[0].get("st_nome_cond")
    return True, nome_condominio


def buscar_unidades(id_condominio: int):
    mapa = {}
    pagina = 1
    while True:
        response = requests.get(
            f"{settings.SUPERLOGICA_BASE_URL}/unidades",
            headers=_get_headers(),
            params={"idCondominio": id_condominio, "pagina": pagina, "itensPorPagina": 50},
            timeout=30,
        )
        if response.status_code != 200:
            return None
        dados = response.json()
        if not dados:
            break
        for unidade in dados:
            unidade_id = unidade.get("id_unidade_uni")
            codigo_unidade = (unidade.get("st_unidade_uni") or "").strip()
            sacado = (
                unidade.get("st_sacado_uni") or unidade.get("nome_proprietario") or ""
            ).strip()
            nome_pdf = f"{codigo_unidade} - {sacado}".strip(" -")
            telefones = []
            for campo in ["telefone_proprietario", "celular_proprietario"]:
                numero = unidade.get(campo)
                if numero and str(numero).strip():
                    telefones.append(str(numero).strip())
            mapa[unidade_id] = {
                "unidade": codigo_unidade,
                "sacado": sacado,
                "nome_pdf": nome_pdf,
                "telefones": list(dict.fromkeys(telefones)),
            }
        pagina += 1
    return mapa


def _formatar_data(valor) -> str:
    if not valor:
        return ""
    texto = str(valor).strip()
    if len(texto) >= 10 and texto[2] == "/" and texto[5] == "/":
        return texto[:10]
    if len(texto) >= 10 and texto[4] == "-" and texto[7] == "-":
        partes = texto[:10].split("-")
        return f"{partes[2]}/{partes[1]}/{partes[0]}"
    return texto


def _para_decimal(valor) -> Decimal:
    if valor in (None, "", "null"):
        return Decimal("0")
    texto = str(valor).strip()
    try:
        return Decimal(texto)
    except InvalidOperation:
        pass
    try:
        return Decimal(texto.replace(".", "").replace(",", "."))
    except InvalidOperation:
        return Decimal("0")


def _d2f(valor: Decimal) -> float:
    return float(valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _descobrir_unidades_inadimplentes(id_condominio: int, data_posicao: str) -> set:
    """
    Usa /index (rápido) para descobrir quais unidades têm inadimplência.
    Retorna um set de id_unidade_uni.
    """
    unidades = set()
    pagina = 1
    processados = set()

    while True:
        response = requests.get(
            f"{settings.SUPERLOGICA_BASE_URL}/inadimplencia/index",
            headers=_get_headers(),
            params={
                "idCondominio":      id_condominio,
                "pagina":            pagina,
                "itensPorPagina":    50,
                "posicaoEm":         data_posicao,
                "comValoresAtualizados": 1,
            },
            timeout=30,
        )
        if response.status_code != 200:
            break
        dados = response.json()
        if not dados:
            break
        for item in dados:
            if not isinstance(item, dict):
                continue
            for receb in item.get("recebimento", []):
                if not isinstance(receb, dict):
                    continue
                if receb.get("fl_status_recb") != "0":
                    continue
                id_r = receb.get("id_recebimento_recb")
                if id_r in processados:
                    continue
                processados.add(id_r)
                unidades.add(receb.get("id_unidade_uni"))
        pagina += 1

    return unidades


def _buscar_valores_unidade(id_condominio: int, id_unidade: str, mapa_unidades: dict):
    """
    Usa /avancada filtrando por unidade para obter valores exatos
    com índice de correção monetária atualizado.
    """
    # Retry até 3 vezes em caso de erro temporário
    for tentativa in range(3):
        try:
            response = requests.get(
                f"{settings.SUPERLOGICA_BASE_URL}/inadimplencia/avancada",
                headers=_get_headers(),
                params={
                    "idCondominio":           id_condominio,
                    "idUnidades":             id_unidade,
                    "itensPorPagina":         500,
                    "comEncargos":            "true",
                    "comHonorarios":          "true",
                    "comAtualizacaoMonetaria": "true",
                },
                timeout=60,
            )
            if response.status_code == 200:
                break
            if tentativa < 2:
                time.sleep(1)
        except requests.RequestException:
            if tentativa < 2:
                time.sleep(1)
            else:
                return None, None
    else:
        return None, None

    if response.status_code != 200:
        return None, None

    dados = response.json()
    if not dados:
        return None, None

    dados_unidade = mapa_unidades.get(id_unidade, {})
    nome_pdf  = dados_unidade.get("nome_pdf", f"Unidade {id_unidade}")
    telefones = dados_unidade.get("telefones", [])

    resumo = {
        "nome_pdf":    nome_pdf,
        "telefones":   telefones,
        "vencimento":  "",
        "competencia": "",
        "principal":   Decimal("0"),
        "juros":       Decimal("0"),
        "multa":       Decimal("0"),
        "atualizacao": Decimal("0"),
        "honorarios":  Decimal("0"),
        "total":       Decimal("0"),
    }
    detalhado = []

    for receb in dados:
        if not isinstance(receb, dict):
            continue

        vencimento  = receb.get("dt_vencimento_recb", "")
        competencia = receb.get("dt_competencia_recb", "")
        id_receb    = receb.get("id_recebimento_recb")

        if not resumo["vencimento"]:
            resumo["vencimento"]  = _formatar_data(vencimento)
            resumo["competencia"] = _formatar_data(competencia)

        det = receb.get("detalhes", {}) or {}
        p = _para_decimal(receb.get("total_receita") or receb.get("vl_emitido_recb") or 0)
        j = _para_decimal(det.get("juros",      0))
        m = _para_decimal(det.get("multa",      0))
        h = _para_decimal(det.get("honorarios", 0))
        a = _para_decimal(det.get("atualizacaomonetaria") or det.get("atualizacao") or 0)
        total_receb = p + j + m + a + h

        resumo["principal"]   += p
        resumo["juros"]       += j
        resumo["multa"]       += m
        resumo["atualizacao"] += a
        resumo["honorarios"]  += h
        resumo["total"]       += total_receb

        detalhado.append({
            "Condomínio":  None,
            "id_unidade":  id_unidade,
            "Unidade":     nome_pdf,
            "Código":      id_receb,
            "Vencimento":  _formatar_data(vencimento),
            "Competência": _formatar_data(competencia),
            "Principal":   _d2f(p),
            "Juros":       _d2f(j),
            "Multa":       _d2f(m),
            "Atualização": _d2f(a),
            "Honorários":  _d2f(h),
            "Total":       _d2f(total_receb),
        })

    # Converte resumo para float
    for campo in ("principal", "juros", "multa", "atualizacao", "honorarios", "total"):
        resumo[campo] = _d2f(resumo[campo])

    return resumo, detalhado


# Número de threads paralelas por condomínio.
# 8 é um bom equilíbrio: rápido sem sobrecarregar a API Superlógica.
_MAX_WORKERS_UNIDADES = 12
# Número de condomínios processados em paralelo.
_MAX_WORKERS_CONDOMINIOS = 6


def buscar_inadimplentes_condominio(id_condominio: int, data_posicao: str, mapa_unidades: dict):
    """
    Estratégia em 2 etapas com paralelismo:
    1. /index  → descobre quais unidades são inadimplentes (rápido)
    2. /avancada por unidade → busca valores exatos em paralelo (rápido + preciso)
    """
    unidades_ids = _descobrir_unidades_inadimplentes(id_condominio, data_posicao)
    if not unidades_ids:
        return {}, []

    resumo_total    = {}
    detalhado_total = []

    # Busca todos os valores em paralelo com pool de threads
    with ThreadPoolExecutor(max_workers=_MAX_WORKERS_UNIDADES) as executor:
        futures = {
            executor.submit(_buscar_valores_unidade, id_condominio, id_uni, mapa_unidades): id_uni
            for id_uni in unidades_ids
        }
        for future in as_completed(futures):
            id_uni = futures[future]
            try:
                resumo_uni, det_uni = future.result()
                if resumo_uni:
                    resumo_total[id_uni] = resumo_uni
                if det_uni:
                    detalhado_total.extend(det_uni)
            except Exception:
                pass  # unidade com erro é ignorada silenciosamente

    return resumo_total, detalhado_total


# Borda fina para grade
_BORDA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_FILL_PAR   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # cinza claro
_FILL_IMPAR = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # branco


# Larguras fixas por nome de coluna (caracteres)
_LARGURAS_FIXAS = {
    "Condomínio":  55,
    "Unidade":     18,
    "Nome":        45,
    "Telefone 1":  22,
    "Telefone 2":  22,
    "Qtd Inadimpl.": 14,
    "Vencimento":  16,
    "Competência": 16,
    "Principal":   16,
    "Juros":       16,
    "Multa":       12,
    "Atualização": 16,
    "Honorários":  16,
    "Total":       16,
    "Código":      14,
}


def _ajustar_larguras(ws):
    """Aplica larguras fixas por coluna e altura padrão nas linhas."""
    # Lê cabeçalhos da linha 1
    headers = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    for idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        largura = _LARGURAS_FIXAS.get(header, 20)
        ws.column_dimensions[col_letter].width = largura
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 30


def _aplicar_grade_e_zebra(ws, skip_rows=0):
    """Aplica borda em todas as células e alterna cores por linha (zebra)."""
    max_row = ws.max_row
    max_col = ws.max_column
    data_start = skip_rows + 2  # linha após o cabeçalho
    header_row = skip_rows + 1

    # Descobre índice da coluna "Qtd Inadimpl." para forçar alinhamento à esquerda
    qtd_col_idx = None
    for col_idx in range(1, max_col + 1):
        if ws.cell(row=header_row, column=col_idx).value == "Qtd Inadimpl.":
            qtd_col_idx = col_idx
            break

    for row_idx in range(data_start, max_row + 1):
        fill = _FILL_PAR if row_idx % 2 == 0 else _FILL_IMPAR
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDA
            if col_idx in (1, 3):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_idx == qtd_col_idx:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_idx < max_row:
                cell.fill = fill
    # Borda também no cabeçalho
    for cell in ws[header_row]:
        cell.border = _BORDA


def _estilizar_cabecalho(ws, header_row=1):
    header_fill = PatternFill(start_color="006837", end_color="006837", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _aplicar_formato_contabil(ws, headers):
    """Aplica formato contábil R$ nas colunas numéricas."""
    cols_numericas = {"Principal", "Juros", "Multa", "Atualização", "Honorários", "Total"}
    fmt = r'R$ #,##0.00'
    for col_idx, header in enumerate(headers, start=1):
        if header in cols_numericas:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt


def _sort_id(uid) -> int:
    """Chave de ordenação numérica pelo id da unidade."""
    try:
        return int(uid)
    except (TypeError, ValueError):
        return 0


def gerar_relatorio_inadimplentes(
    id_condominio: Optional[int] = None,
    data_posicao: Optional[str] = None,
) -> tuple:
    if not data_posicao:
        data_posicao = datetime.today().strftime("%m/%d/%Y")
    else:
        partes = data_posicao.split("/")
        if len(partes) == 3:
            data_posicao = f"{partes[1]}/{partes[0]}/{partes[2]}"

    ids_range = [id_condominio] if id_condominio else range(1, getattr(settings, "SUPERLOGICA_MAX_ID", 100) + 1)

    todas_resumo = []
    todo_detalhado = []

    def _processar_condominio(condo_id):
        """Processa um condomínio completo e retorna (linhas_resumo, linhas_detalhado)."""
        acesso, nome_condominio = verificar_condominio(condo_id)
        if not acesso:
            return [], []
        mapa_unidades = buscar_unidades(condo_id)
        if not mapa_unidades:
            return [], []
        resumo, detalhado = buscar_inadimplentes_condominio(condo_id, data_posicao, mapa_unidades)
        if not resumo:
            return [], []

        # Conta inadimplências por unidade a partir do detalhado
        contagem_por_unidade = {}
        for row_det in detalhado:
            uid = row_det.get("id_unidade")
            nome_uni = row_det.get("Unidade", "")
            chave = uid or nome_uni
            contagem_por_unidade[chave] = contagem_por_unidade.get(chave, 0) + 1

        linhas_resumo = []
        # ── ALTERAÇÃO: itera ordenando pelo id numérico da unidade (menor → maior) ──
        for unidade_id in sorted(resumo.keys(), key=_sort_id):
            valores = resumo[unidade_id]
            telefones = valores.get("telefones", [])
            dados_uni = mapa_unidades.get(unidade_id, {})
            tel1 = telefones[0] if len(telefones) > 0 else "s/n"
            tel2 = telefones[1] if len(telefones) > 1 else "s/n"
            nome_uni = dados_uni.get("unidade") or valores["nome_pdf"]
            qtd = contagem_por_unidade.get(unidade_id) or contagem_por_unidade.get(nome_uni, 0)
            linhas_resumo.append({
                "Condomínio":       nome_condominio,
                "Unidade":          nome_uni,
                "Nome":             dados_uni.get("sacado", ""),
                "Telefone 1":       tel1,
                "Telefone 2":       tel2,
                "Qtd Inadimpl.":    qtd,
                "Vencimento":       valores.get("vencimento", ""),
                "Competência":      valores.get("competencia", ""),
                "Principal":        valores["principal"],
                "Juros":            valores["juros"],
                "Multa":            valores["multa"],
                "Atualização":      valores["atualizacao"],
                "Honorários":       valores["honorarios"],
                "Total":            valores["total"],
            })

        for row in detalhado:
            row["Condomínio"] = nome_condominio

        return linhas_resumo, detalhado

    # Processa condomínios em paralelo
    with ThreadPoolExecutor(max_workers=_MAX_WORKERS_CONDOMINIOS) as executor:
        futures = {executor.submit(_processar_condominio, cid): cid for cid in ids_range}
        for future in as_completed(futures):
            try:
                linhas_r, linhas_d = future.result()
                todas_resumo.extend(linhas_r)
                todo_detalhado.extend(linhas_d)
            except Exception:
                pass

    if not todas_resumo:
        return None, None

    # ── ALTERAÇÃO: ordena por condomínio (alfabético) depois por id numérico ──
    todas_resumo.sort(key=lambda r: (r["Condomínio"] or "", _sort_id(
        next((k for k, v in {}.items()), r["Unidade"])  # fallback ao nome se id perdido
    )))
    # O detalhado já carrega id_unidade — ordena por ele
    todo_detalhado.sort(key=lambda r: (r["Condomínio"] or "", _sort_id(r.get("id_unidade"))))

    wb = Workbook()

    # ── Aba Resumo ──────────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    headers_resumo = ["Condomínio", "Unidade", "Nome", "Telefone 1", "Telefone 2", "Qtd Inadimpl.",
                      "Vencimento", "Competência",
                      "Principal", "Juros", "Multa", "Atualização", "Honorários", "Total"]
    ws_resumo.append(headers_resumo)
    for row in todas_resumo:
        ws_resumo.append([row[h] for h in headers_resumo])

    # Linha de totais
    num_rows = len(todas_resumo)
    totais_resumo = ["TOTAL GERAL", "", "", "", "", "", "", ""]
    cols_num = ["Principal", "Juros", "Multa", "Atualização", "Honorários", "Total"]
    for col in cols_num:
        totais_resumo.append(round(sum(r[col] for r in todas_resumo), 2))
    ws_resumo.append(totais_resumo)

    # Estilo da linha de totais
    total_row_idx = num_rows + 2
    total_fill = PatternFill(start_color="004225", end_color="004225", fill_type="solid")
    total_font = Font(color="FFFFFF", bold=True)
    for cell in ws_resumo[total_row_idx]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _estilizar_cabecalho(ws_resumo)
    _ajustar_larguras(ws_resumo)
    _aplicar_grade_e_zebra(ws_resumo)
    _aplicar_formato_contabil(ws_resumo, headers_resumo)

    # ── Aba Detalhado ────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detalhado")
    headers_det = ["Condomínio", "Unidade", "Código", "Vencimento", "Competência",
                   "Principal", "Juros", "Multa", "Atualização", "Honorários", "Total"]
    ws_det.append(headers_det)
    for row in todo_detalhado:
        ws_det.append([row.get(h, "") for h in headers_det])

    # Linha de totais detalhado
    num_rows_det = len(todo_detalhado)
    totais_det = ["TOTAL GERAL", "", "", "", ""]
    cols_num_det = ["Principal", "Juros", "Multa", "Atualização", "Honorários", "Total"]
    for col in cols_num_det:
        totais_det.append(round(sum(r.get(col, 0) for r in todo_detalhado), 2))
    ws_det.append(totais_det)

    total_row_det = num_rows_det + 2
    for cell in ws_det[total_row_det]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _estilizar_cabecalho(ws_det)
    _ajustar_larguras(ws_det)
    _aplicar_grade_e_zebra(ws_det)
    _aplicar_formato_contabil(ws_det, headers_det)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_condo{id_condominio}" if id_condominio else "_todos"
    filename = f"inadimplentes{suffix}_{timestamp}.xlsx"

    return buffer.getvalue(), filename


# ── Geração de PDF ────────────────────────────────────────────────────────────

def gerar_pdf_inadimplentes(
    id_condominio: Optional[int] = None,
    data_posicao: Optional[str] = None,
) -> tuple:
    """
    Gera um relatório PDF de inadimplentes com tabela formatada.
    Reutiliza o mesmo pipeline de dados do Excel.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, PageBreak, HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    # Cores padrão do sistema
    COR_VERDE       = colors.HexColor("#006837")
    COR_VERDE_ESC   = colors.HexColor("#004225")
    COR_ZEBRA       = colors.HexColor("#F2F2F2")
    COR_BRANCO      = colors.white
    COR_TEXTO       = colors.HexColor("#1A1A1A")
    COR_TEXTO_CLARO = colors.HexColor("#666666")

    # ── Busca os dados ────────────────────────────────────────────────────────
    if not data_posicao:
        data_posicao_fmt = datetime.today().strftime("%m/%d/%Y")
    else:
        partes = data_posicao.split("/")
        if len(partes) == 3:
            data_posicao_fmt = f"{partes[1]}/{partes[0]}/{partes[2]}"
        else:
            data_posicao_fmt = data_posicao

    ids_range = [id_condominio] if id_condominio else range(1, getattr(settings, "SUPERLOGICA_MAX_ID", 100) + 1)

    todas_resumo = []

    def _processar_condo_pdf(condo_id):
        acesso, nome_condo = verificar_condominio(condo_id)
        if not acesso:
            return []
        mapa = buscar_unidades(condo_id)
        if not mapa:
            return []
        resumo, _ = buscar_inadimplentes_condominio(condo_id, data_posicao_fmt, mapa)
        if not resumo:
            return []
        linhas = []
        # ── ALTERAÇÃO: itera ordenando pelo id numérico da unidade (menor → maior) ──
        for uid in sorted(resumo.keys(), key=_sort_id):
            vals = resumo[uid]
            tels = vals.get("telefones", [])
            dados_uni = mapa.get(uid, {})
            linhas.append({
                "Condomínio":   nome_condo or "",
                "condo_id":     condo_id,
                "Unidade":      dados_uni.get("unidade") or vals["nome_pdf"],
                "Nome":         dados_uni.get("sacado", ""),
                "Telefone 1":   tels[0] if len(tels) > 0 else "s/n",
                "Telefone 2":   tels[1] if len(tels) > 1 else "s/n",
                "Principal":    vals["principal"],
                "Juros":        vals["juros"],
                "Multa":        vals["multa"],
                "Atualização":  vals["atualizacao"],
                "Honorários":   vals["honorarios"],
                "Total":        vals["total"],
            })
        return linhas

    with ThreadPoolExecutor(max_workers=_MAX_WORKERS_CONDOMINIOS) as executor:
        futures = {executor.submit(_processar_condo_pdf, cid): cid for cid in ids_range}
        for future in as_completed(futures):
            try:
                todas_resumo.extend(future.result())
            except Exception:
                pass

    if not todas_resumo:
        return None, None

    # ── ALTERAÇÃO: ordena por condomínio (alfabético) depois por id numérico ──
    todas_resumo.sort(key=lambda r: (r["Condomínio"], r["Unidade"]))

    # ── Monta o PDF ───────────────────────────────────────────────────────────
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,    bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        "titulo", fontSize=16, textColor=COR_VERDE,
        fontName="Helvetica-Bold", alignment=TA_LEFT, spaceAfter=4,
    )
    style_sub = ParagraphStyle(
        "sub", fontSize=9, textColor=COR_TEXTO_CLARO,
        fontName="Helvetica", alignment=TA_LEFT, spaceAfter=2,
    )
    style_cell = ParagraphStyle(
        "cell", fontSize=7, textColor=COR_TEXTO,
        fontName="Helvetica", leading=9,
    )
    style_cell_bold = ParagraphStyle(
        "cell_bold", fontSize=7, textColor=COR_BRANCO,
        fontName="Helvetica-Bold", leading=9, alignment=TA_CENTER,
    )

    def brl(valor):
        """Formata Decimal como R$ 1.234,56"""
        try:
            v = float(valor)
            return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return str(valor)

    def p(txt, style=None):
        return Paragraph(str(txt) if txt else "", style or style_cell)

    story = []

    # Título
    data_label = data_posicao or datetime.today().strftime("%d/%m/%Y")
    titulo_txt = "Relatório de Inadimplentes"
    if id_condominio:
        titulo_txt += f" — Condomínio {id_condominio}"
    # Logo + Título lado a lado
    import os as _os
    from reportlab.platypus import Image as RLImage
    _logo_path = _os.path.join(_os.path.dirname(__file__), "logo_pratika.png")

    if _os.path.exists(_logo_path):
        logo_img = RLImage(_logo_path, width=3.5*cm, height=1.8*cm)
    else:
        logo_img = Paragraph("", style_sub)

    header_table = Table(
        [[logo_img, Paragraph(titulo_txt, style_title)]],
        colWidths=[4*cm, None],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_table)
    story.append(Paragraph(f"Data de posição: {data_label}  |  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total de unidades: {len(todas_resumo)}", style_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=COR_VERDE, spaceAfter=10))

    # Acumuladores do total geral
    grand_principal = grand_juros = grand_multa = grand_atualiz = grand_honor = grand_total = Decimal("0")

    # Agrupa por condomínio para uma tabela por condomínio
    from itertools import groupby
    for (nome_condo, cid), grupo in groupby(todas_resumo, key=lambda r: (r["Condomínio"], r["condo_id"])):
        rows = list(grupo)

        story.append(Paragraph(
            f"[ID {cid}] {nome_condo or 'Sem nome'}",
            ParagraphStyle(
                "condo_header", fontSize=10, textColor=COR_VERDE,
                fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=10,
            )
        ))

        # Cabeçalho da tabela
        cabecalho = [
            p("Unidade", style_cell_bold),
            p("Nome", style_cell_bold),
            p("Telefone 1", style_cell_bold),
            p("Telefone 2", style_cell_bold),
            p("Principal", style_cell_bold),
            p("Juros", style_cell_bold),
            p("Multa", style_cell_bold),
            p("Atualização", style_cell_bold),
            p("Honorários", style_cell_bold),
            p("Total", style_cell_bold),
        ]

        dados_tabela = [cabecalho]
        tot_principal = tot_juros = tot_multa = tot_atualiz = tot_honor = tot_total = Decimal("0")

        for i, row in enumerate(rows):
            dados_tabela.append([
                p(row["Unidade"]),
                p(row["Nome"]),
                p(row["Telefone 1"]),
                p(row["Telefone 2"]),
                p(brl(row["Principal"])),
                p(brl(row["Juros"])),
                p(brl(row["Multa"])),
                p(brl(row["Atualização"])),
                p(brl(row["Honorários"])),
                p(brl(row["Total"])),
            ])
            tot_principal += Decimal(str(row["Principal"]))
            tot_juros     += Decimal(str(row["Juros"]))
            tot_multa     += Decimal(str(row["Multa"]))
            tot_atualiz   += Decimal(str(row["Atualização"]))
            tot_honor     += Decimal(str(row["Honorários"]))
            tot_total     += Decimal(str(row["Total"]))
            grand_principal += Decimal(str(row["Principal"]))
            grand_juros     += Decimal(str(row["Juros"]))
            grand_multa     += Decimal(str(row["Multa"]))
            grand_atualiz   += Decimal(str(row["Atualização"]))
            grand_honor     += Decimal(str(row["Honorários"]))
            grand_total     += Decimal(str(row["Total"]))

        # Linha de totais
        style_tot = ParagraphStyle("tot", fontSize=7, textColor=COR_BRANCO,
                                   fontName="Helvetica-Bold", alignment=TA_CENTER)
        dados_tabela.append([
            p("TOTAL", style_tot),
            p("", style_tot),
            p("", style_tot),
            p("", style_tot),
            p(brl(tot_principal), style_tot),
            p(brl(tot_juros), style_tot),
            p(brl(tot_multa), style_tot),
            p(brl(tot_atualiz), style_tot),
            p(brl(tot_honor), style_tot),
            p(brl(tot_total), style_tot),
        ])

        # Larguras das colunas (total ~26cm em landscape A4)
        col_widths = [2.5*cm, 5*cm, 3*cm, 3*cm, 2.5*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm, 3*cm]

        table_style = [
            # Cabeçalho
            ("BACKGROUND",  (0, 0), (-1, 0),  COR_VERDE),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  COR_BRANCO),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  7),
            ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUND", (0, 1), (-1, -2),
             [COR_BRANCO if i % 2 == 0 else COR_ZEBRA for i in range(len(rows))]),
            # Totais
            ("BACKGROUND",  (0, -1), (-1, -1), COR_VERDE_ESC),
            ("TEXTCOLOR",   (0, -1), (-1, -1), COR_BRANCO),
            # Grade
            ("GRID",        (0, 0), (-1, -1),  0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0, 0), (-1, -1),  4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1),  4),
            ("RIGHTPADDING",(0, 0), (-1, -1),  4),
        ]

        tabela = Table(dados_tabela, colWidths=col_widths, repeatRows=1)
        tabela.setStyle(TableStyle(table_style))
        story.append(tabela)
        story.append(Spacer(1, 0.3*cm))

    # ── Tabela de TOTAL GERAL (todos os condomínios) ─────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=COR_VERDE_ESC, spaceAfter=6))
    story.append(Paragraph("TOTAL GERAL — Todos os Condomínios", ParagraphStyle(
        "grand_title", fontSize=11, textColor=COR_VERDE_ESC,
        fontName="Helvetica-Bold", spaceAfter=6,
    )))

    style_grand = ParagraphStyle("grand", fontSize=8, textColor=COR_BRANCO,
                                  fontName="Helvetica-Bold", alignment=TA_CENTER)
    style_grand_lbl = ParagraphStyle("grand_lbl", fontSize=8, textColor=COR_BRANCO,
                                      fontName="Helvetica-Bold", alignment=TA_LEFT)

    dados_grand = [
        [
            Paragraph("Principal", style_grand),
            Paragraph("Juros", style_grand),
            Paragraph("Multa", style_grand),
            Paragraph("Atualização", style_grand),
            Paragraph("Honorários", style_grand),
            Paragraph("TOTAL GERAL", style_grand),
        ],
        [
            Paragraph(brl(grand_principal), style_grand),
            Paragraph(brl(grand_juros), style_grand),
            Paragraph(brl(grand_multa), style_grand),
            Paragraph(brl(grand_atualiz), style_grand),
            Paragraph(brl(grand_honor), style_grand),
            Paragraph(brl(grand_total), style_grand),
        ],
    ]

    tabela_grand = Table(
        dados_grand,
        colWidths=[4*cm, 3.5*cm, 3.5*cm, 4*cm, 4*cm, 5*cm],
    )
    tabela_grand.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  COR_VERDE),
        ("BACKGROUND",   (0, 1), (-1, 1),  COR_VERDE_ESC),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tabela_grand)

    doc.build(story)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_condo{id_condominio}" if id_condominio else "_todos"
    filename = f"inadimplentes{suffix}_{timestamp}.pdf"

    return buffer.getvalue(), filename